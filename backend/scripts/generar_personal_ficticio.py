"""Genera archivos Excel ficticios de Jurados y Testigos para pruebas.

Uso:
    python scripts/generar_personal_ficticio.py [--puestos N] [--jurados-por-puesto N] [--testigos-por-puesto N]

Produce:
    data/ficticio_jurados.xlsx
    data/ficticio_testigos.xlsx

Los archivos tienen la misma estructura que las plantillas originales (sheet "Formulario").
"""
from __future__ import annotations

import argparse
import random
import sys
from pathlib import Path

# Add backend root to sys.path so app.* imports work when run from project root
sys.path.insert(0, str(Path(__file__).parent.parent))

import openpyxl
from sqlalchemy.orm import Session

from app.database import SessionLocal
from app.db_models import PuestoORM


# ── Datos ficticios ─────────────────────────────────────────────────────────

NOMBRES = [
    "ANA", "CARLOS", "DIANA", "EDGAR", "FERNANDA", "GUILLERMO", "HELENA",
    "IVAN", "JULIANA", "KEVIN", "LAURA", "MARIO", "NATALIA", "OSCAR",
    "PAOLA", "RICARDO", "SANDRA", "TOMAS", "URSULA", "VICTOR",
    "WENDY", "XAVIER", "YOLANDA", "ZEUS", "ALBA", "BERNARDO", "CAMILO",
    "DOLORES", "EMILIO", "FABIOLA",
]

APELLIDOS = [
    "GARCIA", "MARTINEZ", "RODRIGUEZ", "LOPEZ", "GONZALEZ", "PEREZ",
    "SANCHEZ", "RAMIREZ", "TORRES", "FLORES", "VARGAS", "MORALES",
    "JIMENEZ", "CASTRO", "ROMERO", "GUTIERREZ", "DIAZ", "REYES",
    "HERRERA", "MENDEZ", "SILVA", "RAMOS", "SUAREZ", "MOLINA",
    "SALAZAR", "ORTIZ", "CARDONA", "VELASQUEZ", "MUNOZ", "RIOS",
]

NIVELES_EDUCATIVOS = [
    "BACHILLER", "TECNICO", "TECNOLOGO", "UNIVERSITARIO", "POSGRADO", "PRIMARIA"
]

REFERENCIADOS = [
    "PARTIDO LIBERAL", "PARTIDO CONSERVADOR", "CAMBIO RADICAL",
    "COLOMBIA HUMANA", "CENTRO DEMOCRATICO", "ALIANZA VERDE", "NINGUNO",
]


def _cedula() -> str:
    return str(random.randint(10_000_000, 1_099_999_999))


def _telefono() -> str:
    return f"3{random.randint(100_000_000, 199_999_999)}"


def _correo(nombre: str, apellido: str) -> str:
    n = nombre.lower()
    a = apellido.lower()
    domains = ["gmail.com", "hotmail.com", "yahoo.com", "outlook.com"]
    return f"{n}.{a}{random.randint(1,99)}@{random.choice(domains)}"


def _persona():
    nombre1 = random.choice(NOMBRES)
    nombre2 = random.choice(NOMBRES) if random.random() > 0.4 else ""
    apellido1 = random.choice(APELLIDOS)
    apellido2 = random.choice(APELLIDOS) if random.random() > 0.3 else ""
    cedula = _cedula()
    tel = _telefono() if random.random() > 0.3 else ""
    cel = _telefono()
    correo = _correo(nombre1, apellido1) if random.random() > 0.2 else ""
    return {
        "nombre1": nombre1, "nombre2": nombre2,
        "apellido1": apellido1, "apellido2": apellido2,
        "cedula": cedula, "tel": tel, "cel": cel, "correo": correo,
    }


# ── Excel writers ────────────────────────────────────────────────────────────

JURADOS_HEADERS = [
    "DEPARTAMENTO", "MUNICIPIO", "PUESTO", "CEDULA", "PRIMER NOMBRE", "SEGUNDO NOMBRE",
    "PRIMER APELLIDO", "SEGUNDO APELLIDO", "DIRECCION", "TELEFONO", "CELULAR",
    "CORREO", "NIVEL EDUCATIVO", "REFERENCIADO POR",
]

JURADOS_HEADERS_CON_CODIGO = [
    "DEPARTAMENTO", "MUNICIPIO", "PUESTO", "CODIGO", "CEDULA", "PRIMER NOMBRE", "SEGUNDO NOMBRE",
    "PRIMER APELLIDO", "SEGUNDO APELLIDO", "DIRECCION", "TELEFONO", "CELULAR",
    "CORREO", "NIVEL EDUCATIVO", "REFERENCIADO POR",
]

TESTIGOS_HEADERS = [
    "DEPARTAMENTO", "MUNICIPIO", "CEDULA", "PRIMER NOMBRE", "SEGUNDO NOMBRE",
    "PRIMER APELLIDO", "SEGUNDO APELLIDO", "CEDULA", "TELEFONO", "CORREO",
    "CELULAR", "COMUNA/LOCALIDAD/ZONA", "PUESTO DE VOTACION OPCION 1", "REFERENCIADO POR",
]

TESTIGOS_HEADERS_CON_CODIGO = [
    "DEPARTAMENTO", "MUNICIPIO", "CODIGO", "CEDULA", "PRIMER NOMBRE", "SEGUNDO NOMBRE",
    "PRIMER APELLIDO", "SEGUNDO APELLIDO", "CEDULA", "TELEFONO", "CORREO",
    "CELULAR", "COMUNA/LOCALIDAD/ZONA", "PUESTO DE VOTACION OPCION 1", "REFERENCIADO POR",
]


def _corrupt_name(name: str) -> str:
    """Introduce a small typo in a name to simulate real-world mismatches."""
    if not name or len(name) < 4:
        return name
    choices = [
        lambda s: s.replace("A", "4", 1),          # digit substitution
        lambda s: s[:-2] + s[-1],                   # drop last char
        lambda s: s[:3] + "." + s[3:],              # insert punct
        lambda s: s.upper().replace("O", "0", 1),   # o→0
        lambda s: s + " " + s[-2:],                 # trailing junk
    ]
    return random.choice(choices)(name)


def _write_jurados(path: Path, rows: list[dict]) -> None:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Formulario"
    ws.append(JURADOS_HEADERS)
    for r in rows:
        ws.append([
            r["departamento"],
            r["municipio"],
            r["puesto"],
            r["cedula"],
            r["nombre1"],
            r["nombre2"],
            r["apellido1"],
            r["apellido2"],
            f"CRA {random.randint(1,100)} # {random.randint(1,50)}-{random.randint(1,99)}",
            r["tel"],
            r["cel"],
            r["correo"],
            random.choice(NIVELES_EDUCATIVOS),
            random.choice(REFERENCIADOS),
        ])
    wb.save(path)
    print(f"  Jurados → {path} ({len(rows)} filas)")


def _write_jurados_con_codigo(path: Path, rows: list[dict]) -> None:
    """Igual que _write_jurados pero con columna CODIGO y nombres de puesto con errores."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Formulario"
    ws.append(JURADOS_HEADERS_CON_CODIGO)
    for r in rows:
        # ~30% de filas tienen nombre de puesto con error tipográfico → debe resolverse por CODIGO
        puesto_nombre = _corrupt_name(r["puesto"]) if random.random() < 0.3 else r["puesto"]
        ws.append([
            r["departamento"],
            r["municipio"],
            puesto_nombre,
            r["codigo_puesto"],   # CODIGO — resuelve aunque el nombre esté mal
            r["cedula"],
            r["nombre1"],
            r["nombre2"],
            r["apellido1"],
            r["apellido2"],
            f"CRA {random.randint(1,100)} # {random.randint(1,50)}-{random.randint(1,99)}",
            r["tel"],
            r["cel"],
            r["correo"],
            random.choice(NIVELES_EDUCATIVOS),
            random.choice(REFERENCIADOS),
        ])
    wb.save(path)
    print(f"  Jurados (con CODIGO) → {path} ({len(rows)} filas)")


def _write_testigos(path: Path, rows: list[dict]) -> None:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Formulario"
    ws.append(TESTIGOS_HEADERS)
    for r in rows:
        ws.append([
            r["departamento"],
            r["municipio"],
            r["cedula"],
            r["nombre1"],
            r["nombre2"],
            r["apellido1"],
            r["apellido2"],
            r["cedula"],   # duplicated CEDULA column (as in original template)
            r["tel"],
            r["correo"],
            r["cel"],
            r["comuna"],
            r["puesto"],
            random.choice(REFERENCIADOS),
        ])
    wb.save(path)
    print(f"  Testigos → {path} ({len(rows)} filas)")


def _write_testigos_con_codigo(path: Path, rows: list[dict]) -> None:
    """Igual que _write_testigos pero con columna CODIGO y nombres con errores."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Formulario"
    ws.append(TESTIGOS_HEADERS_CON_CODIGO)
    for r in rows:
        puesto_nombre = _corrupt_name(r["puesto"]) if random.random() < 0.3 else r["puesto"]
        ws.append([
            r["departamento"],
            r["municipio"],
            r["codigo_puesto"],   # CODIGO
            r["cedula"],
            r["nombre1"],
            r["nombre2"],
            r["apellido1"],
            r["apellido2"],
            r["cedula"],
            r["tel"],
            r["correo"],
            r["cel"],
            r["comuna"],
            puesto_nombre,        # nombre con posible error tipográfico (ignorado porque hay CODIGO)
            random.choice(REFERENCIADOS),
        ])
    wb.save(path)
    print(f"  Testigos (con CODIGO) → {path} ({len(rows)} filas)")


# ── Main ─────────────────────────────────────────────────────────────────────

def main() -> None:
    parser = argparse.ArgumentParser(description="Genera datos ficticios de jurados y testigos.")
    parser.add_argument("--puestos", type=int, default=200,
                        help="Número de puestos a seleccionar aleatoriamente (default: 200)")
    parser.add_argument("--jurados-por-puesto", type=int, default=4,
                        help="Jurados por puesto (default: 4)")
    parser.add_argument("--testigos-por-puesto", type=int, default=2,
                        help="Testigos por puesto (default: 2)")
    args = parser.parse_args()

    db: Session = SessionLocal()
    try:
        total_puestos = db.query(PuestoORM).count()
        if total_puestos == 0:
            print("ERROR: No hay puestos en la base de datos. Ejecute primero el script de importación.")
            return

        n = min(args.puestos, total_puestos)
        print(f"Seleccionando {n} puestos de {total_puestos} disponibles…")

        all_puestos = db.query(
            PuestoORM.codigo_puesto,
            PuestoORM.departamento,
            PuestoORM.municipio,
            PuestoORM.puesto,
            PuestoORM.comuna,
        ).all()

        selected = random.sample(all_puestos, n)
    finally:
        db.close()

    jurados_rows: list[dict] = []
    testigos_rows: list[dict] = []

    cedulas_usadas: set[str] = set()

    def unique_cedula() -> str:
        while True:
            c = _cedula()
            if c not in cedulas_usadas:
                cedulas_usadas.add(c)
                return c

    for puesto_row in selected:
        dep = puesto_row.departamento
        mun = puesto_row.municipio
        puesto_name = puesto_row.puesto
        codigo_puesto = puesto_row.codigo_puesto
        comuna = puesto_row.comuna or ""

        for _ in range(args.jurados_por_puesto):
            p = _persona()
            p["cedula"] = unique_cedula()
            jurados_rows.append({**p, "departamento": dep, "municipio": mun, "puesto": puesto_name, "codigo_puesto": codigo_puesto, "comuna": comuna})

        for _ in range(args.testigos_por_puesto):
            p = _persona()
            p["cedula"] = unique_cedula()
            testigos_rows.append({**p, "departamento": dep, "municipio": mun, "puesto": puesto_name, "codigo_puesto": codigo_puesto, "comuna": comuna})

    out_dir = Path(__file__).parent.parent.parent / "data"
    out_dir.mkdir(exist_ok=True)

    _write_jurados(out_dir / "ficticio_jurados.xlsx", jurados_rows)
    _write_testigos(out_dir / "ficticio_testigos.xlsx", testigos_rows)
    _write_jurados_con_codigo(out_dir / "ficticio_jurados_con_codigo.xlsx", jurados_rows)
    _write_testigos_con_codigo(out_dir / "ficticio_testigos_con_codigo.xlsx", testigos_rows)

    print(f"\nListo. Total generado: {len(jurados_rows)} jurados, {len(testigos_rows)} testigos.")
    print("Archivos sin CODIGO: ficticio_jurados.xlsx, ficticio_testigos.xlsx")
    print("Archivos con CODIGO: ficticio_jurados_con_codigo.xlsx, ficticio_testigos_con_codigo.xlsx")
    print("  (los archivos con CODIGO tienen ~30% de nombres de puesto con errores tipográficos)")
    print("Cargue los archivos desde la vista 'JURADOS Y TESTIGOS' en la aplicación.")


if __name__ == "__main__":
    main()
