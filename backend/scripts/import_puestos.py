#!/usr/bin/env python3
"""Script para importar puestos electorales desde Excel a la base de datos."""
import argparse
import random
import sys
from pathlib import Path
from typing import Dict, List, Optional, Tuple

import openpyxl

# Agregar parent dir al path para importar módulos de app
sys.path.append(str(Path(__file__).parent.parent))

from app.database import SessionLocal
from app.db_models import PuestoORM
from sqlalchemy.dialects.postgresql import insert as pg_insert

# Mapeo de variantes de nombres de columnas a nombres estándar
COLUMN_MAPPINGS = {
    'codigo_puesto': ['Codigo_Puesto', 'ID_Puesto', 'CODIGO', 'COD_PUESTO', 'CODIGO PUESTO', 'Cod unico'],
    'departamento_codigo': ['Departamento_Codigo', 'COD_DPTO', 'DPTO_CODIGO', 'Codigo_Departamento', 'COD DPTO', 'dd'],
    'municipio_codigo': ['Municipio_Codigo', 'COD_MPIO', 'MPIO_CODIGO', 'Codigo_Municipio', 'DIVIPOLA', 'COD MUNICIPIO', 'mm'],
    'departamento': ['Departamento', 'DPTO', 'NOMBRE_DPTO', 'DEPARTAMENTO', 'departamento'],
    'municipio': ['Municipio', 'MPIO', 'NOMBRE_MPIO', 'MUNICIPIO', 'municipio'],
    'puesto': ['Puesto', 'Nombre_Puesto', 'NOMBRE', 'NOMBRE_PUESTO', 'PUESTO', 'NOMBRE PUESTO', 'puesto'],
    'direccion': ['Direccion', 'DIRECCION', 'DIR', 'DIRECCIÓN', 'dirección'],
    'comuna': ['Comuna', 'COMUNA', 'Zona', 'ZONA', 'comuna', 'zz'],
    'mesas': ['Mesas', 'MESAS', 'Num_Mesas', 'NÚMERO DE MESAS', 'N° DE MESAS', 'mesas'],
    'mujeres': ['Mujeres', 'MUJERES', 'Potencial_Mujeres', 'POT_MUJERES', 'POTENCIAL MUJERES', 'mujeres'],
    'hombres': ['Hombres', 'HOMBRES', 'Potencial_Hombres', 'POT_HOMBRES', 'POTENCIAL HOMBRES', 'hombres'],
    'total': ['Total', 'TOTAL', 'Potencial_Total', 'POT_TOTAL', 'Potencial', 'POTENCIAL TOTAL', 'total'],
    'latitud': ['Latitud', 'LATITUD', 'LAT', 'Lat', 'LATITUD DECIMAL', 'latitud'],
    'longitud': ['Longitud', 'LONGITUD', 'LON', 'Long', 'LNG', 'LONGITUD DECIMAL', 'longitud'],
    'zona': ['zz', 'Zona', 'ZONA'],
    'puesto_codigo': ['pp'],
}


def detect_columns(header_row: List) -> Dict[str, int]:
    """Detecta índices de columnas basándose en nombres posibles."""
    detected = {}
    for field, variants in COLUMN_MAPPINGS.items():
        for idx, cell_value in enumerate(header_row):
            if cell_value and str(cell_value).strip() in variants:
                detected[field] = idx
                break
    return detected


def validate_row(row_data: Dict, row_num: int) -> Tuple[bool, Optional[str]]:
    """Valida campos obligatorios y formatos."""
    errors = []
    
    # Campos obligatorios
    required = ['codigo_puesto', 'puesto']
    for field in required:
        value = row_data.get(field)
        if not value or (isinstance(value, str) and not value.strip()):
            errors.append(f"Falta campo obligatorio: {field}")
    
    # Validar formato código municipio (5 dígitos) si está presente
    municipio_cod = str(row_data.get('municipio_codigo', '')).strip()
    if municipio_cod:
        # Limpiar y normalizar
        municipio_cod = municipio_cod.replace('.0', '').zfill(5)
        if not municipio_cod.isdigit() or len(municipio_cod) != 5:
            errors.append(f"Código municipio debe ser 5 dígitos numéricos: {municipio_cod}")
    
    # Validar coordenadas si están presentes
    lat = row_data.get('latitud')
    lon = row_data.get('longitud')
    if lat is not None and lat != '':
        try:
            lat_float = float(lat)
            if not (-5 <= lat_float <= 14):
                errors.append(f"Latitud fuera de rango Colombia: {lat_float}")
        except (ValueError, TypeError):
            errors.append(f"Latitud inválida: {lat}")
            
    if lon is not None and lon != '':
        try:
            lon_float = float(lon)
            if not (-82 <= lon_float <= -66):
                errors.append(f"Longitud fuera de rango Colombia: {lon_float}")
        except (ValueError, TypeError):
            errors.append(f"Longitud inválida: {lon}")
    
    # Validar números positivos
    for field in ['mesas', 'mujeres', 'hombres', 'total']:
        val = row_data.get(field)
        if val is not None and val != '':
            try:
                num_val = int(val) if isinstance(val, (int, float, str)) else val
                if num_val < 0:
                    errors.append(f"{field} no puede ser negativo: {num_val}")
            except (ValueError, TypeError):
                pass  # Ignorar si no es número
    
    if errors:
        return False, f"Fila {row_num}: " + "; ".join(errors)
    return True, None


def geocode_fallback(municipio_codigo: str) -> Tuple[float, float]:
    """
    Genera coordenadas aproximadas para un municipio.
    En una implementación completa, esto usaría polígonos reales.
    Por ahora usa el centroide de Colombia con un offset aleatorio.
    """
    # Centroide de Colombia
    base_lat = 4.5709
    base_lon = -74.2973
    
    # Offset aleatorio pequeño (aprox +/- 0.5 grados)
    # En producción, esto debería usar los polígonos reales de municipios
    lat = base_lat + random.uniform(-0.5, 0.5)
    lon = base_lon + random.uniform(-0.5, 0.5)
    
    return lat, lon


def import_puestos(
    file_path: str,
    anio: Optional[int] = None,
    corporacion: Optional[str] = None,
    dry_run: bool = False
) -> Dict:
    """Importa puestos desde Excel a base de datos."""
    
    if not Path(file_path).exists():
        return {'error': f"Archivo no encontrado: {file_path}"}
    
    try:
        wb = openpyxl.load_workbook(file_path, read_only=True, data_only=True)
        ws = wb.active
    except Exception as e:
        return {'error': f"Error leyendo archivo Excel: {e}"}
    
    # Buscar fila de header (buscar en las primeras 10 filas)
    header_row_idx = 1
    header = None
    for row_idx in range(1, min(11, ws.max_row + 1)):
        row_values = [cell.value for cell in ws[row_idx]]
        # Buscar fila que contenga columnas conocidas
        if any(val in str(row_values) for val in ['puesto', 'Puesto', 'mujeres', 'hombres', 'Cod unico']):
            header = row_values
            header_row_idx = row_idx
            break
    
    if header is None:
        return {'error': 'No se encontró fila de encabezados en las primeras 10 filas'}
    
    # Detectar columnas
    col_map = detect_columns(header)
    
    required_cols = ['codigo_puesto', 'puesto']
    missing = [c for c in required_cols if c not in col_map]
    if missing:
        return {'error': f"Columnas faltantes: {missing}. Columnas detectadas: {list(col_map.keys())}"}
    
    # Determinar si hay coordenadasheader_row_idx + 1), start=header_row_idx + 1
    has_coords = 'latitud' in col_map and 'longitud' in col_map
    
    stats = {
        'total_rows': 0,
        'valid': 0,
        'invalid': 0,
        'geocoded': 0,
        'errors': []
    }
    
    puestos_to_insert = []
    
    for row_idx, row in enumerate(ws.iter_rows(min_row=2), start=2):
        stats['total_rows'] += 1
        
        # Extraer datos
        row_data = {}
        for field, col_idx in col_map.items():
            cell_value = row[col_idx].value
            row_data[field] = cell_value
        
        # Normalizar códigos
        # Si tenemos codigo_unico, usarlo directamente
        codigo_unico = row_data.get('codigo_puesto', '')
        if codigo_unico and len(str(codigo_unico)) >= 5:
            codigo_str = str(codigo_unico).replace('.0', '').strip()
            if len(codigo_str) >= 5:
                row_data['municipio_codigo'] = codigo_str[:5]
                row_data['departamento_codigo'] = codigo_str[:2]
                if 'codigo_puesto' in row_data:
                    row_data['codigo_puesto'] = codigo_str
        
        # Si no, construir desde dd + mm
        if not row_data.get('municipio_codigo'):
            dept_cod = str(row_data.get('departamento_codigo', '')).replace('.0', '').strip()
            mun_cod = str(row_data.get('municipio_codigo', '')).replace('.0', '').strip()
            
            if dept_cod and mun_cod:
                row_data['departamento_codigo'] = dept_cod.zfill(2)
                row_data['municipio_codigo'] = dept_cod.zfill(2) + mun_cod.zfill(3)
            elif dept_cod:
                row_data['departamento_codigo'] = dept_cod.zfill(2)
                row_data['municipio_codigo'] = None
        
        # Si no tenemos codigo_puesto, construirlo desde Cod unico o desde componentes
        if not row_data.get('codigo_puesto') or not str(row_data['codigo_puesto']).strip():
            codigo_unico = row_data.get('codigo_puesto', '')
            if codigo_unico:
                row_data['codigo_puesto'] = str(codigo_unico).strip()
            else:
                # Construir desde partes
                parts = [
                    row_data.get('departamento_codigo', ''),
                    row_data.get('municipio_codigo', ''),
                    row_data.get('zona', ''),
                    row_data.get('puesto_codigo', '')
                ]
                row_data['codigo_puesto'] = ''.join(str(p).replace('.0', '').strip() for p in parts if p)
        
        # Validar
        is_valid, error_msg = validate_row(row_data, row_idx)
        if not is_valid:
            stats['invalid'] += 1
            stats['errors'].append(error_msg)
            continue
        
        # Obtener coordenadas
        lat = row_data.get('latitud')
        lon = row_data.get('longitud')
        
        # Convertir a float si es posible
        try:
            if lat is not None and lat != '':
                lat = float(lat)
            else:
                lat = None
        except (ValueError, TypeError):
            lat = None
            
        try:
            if lon is not None and lon != '':
                lon = float(lon)
            else:
                lon = None
        except (ValueError, TypeError):
            lon = None
        
        # Geocodificar si es necesario
        if lat is None or lon is None:
            mpio_code = row_data.get('municipio_codigo')
            if mpio_code:
                lat, lon = geocode_fallback(mpio_code)
                stats['geocoded'] += 1
            else:
                # Sin código de municipio ni coordenadas, usar centroide de Colombia
                lat, lon = 4.5709, -74.2973
                stats['geocoded'] += 1
        
        # Preparar valores numéricos
        def safe_int(val):
            if val is None or val == '':
                return None
            try:
                return int(float(val))
            except (ValueError, TypeError):
                return None
        
        # Preparar registro
        puesto_dict = {
            'codigo_puesto': str(row_data['codigo_puesto']).strip(),
            'departamento_codigo': row_data.get('departamento_codigo', '00'),
            'municipio_codigo': row_data.get('municipio_codigo', '00000'),
            'departamento': str(row_data.get('departamento', '')).strip() or 'N/A',
            'municipio': str(row_data.get('municipio', '')).strip() or 'N/A',
            'puesto': str(row_data['puesto']).strip(),
            'comuna': str(row_data.get('comuna', '')).strip() if row_data.get('comuna') else None,
            'direccion': str(row_data.get('direccion', '')).strip() if row_data.get('direccion') else None,
            'mesas': safe_int(row_data.get('mesas')),
            'mujeres': safe_int(row_data.get('mujeres')),
            'hombres': safe_int(row_data.get('hombres')),
            'total': safe_int(row_data.get('total')),
            'latitud': float(lat),
            'longitud': float(lon),
            'anio': anio,
            'corporacion': corporacion,
        }
        
        puestos_to_insert.append(puesto_dict)
        stats['valid'] += 1
    
    wb.close()
    
    # Insertar en base de datos (upsert)
    if not dry_run and puestos_to_insert:
        db = SessionLocal()
        try:
            # Usar ON CONFLICT para actualizar si ya existe
            stmt = pg_insert(PuestoORM).values(puestos_to_insert)
            stmt = stmt.on_conflict_do_update(
                index_elements=['codigo_puesto'],
                set_={
                    'departamento': stmt.excluded.departamento,
                    'municipio': stmt.excluded.municipio,
                    'puesto': stmt.excluded.puesto,
                    'direccion': stmt.excluded.direccion,
                    'mesas': stmt.excluded.mesas,
                    'mujeres': stmt.excluded.mujeres,
                    'hombres': stmt.excluded.hombres,
                    'total': stmt.excluded.total,
                    'latitud': stmt.excluded.latitud,
                    'longitud': stmt.excluded.longitud,
                    'anio': stmt.excluded.anio,
                    'corporacion': stmt.excluded.corporacion,
                }
            )
            db.execute(stmt)
            db.commit()
            stats['inserted'] = len(puestos_to_insert)
        except Exception as e:
            db.rollback()
            stats['error'] = str(e)
        finally:
            db.close()
    elif dry_run:
        stats['would_insert'] = len(puestos_to_insert)
    
    return stats


def main():
    parser = argparse.ArgumentParser(description='Importar puestos desde Excel')
    parser.add_argument('--file', required=True, help='Ruta al archivo Excel')
    parser.add_argument('--year', type=int, help='Año electoral')
    parser.add_argument('--corporation', choices=['Senado', 'Camara'], help='Corporación')
    parser.add_argument('--dry-run', action='store_true', help='Simular sin escribir DB')
    
    args = parser.parse_args()
    
    print(f"📊 Importando puestos desde: {args.file}")
    if args.dry_run:
        print("🔍 Modo DRY RUN: no se escribirá en la base de datos")
    
    result = import_puestos(
        args.file,
        anio=args.year,
        corporacion=args.corporation,
        dry_run=args.dry_run
    )
    
    if 'error' in result:
        print(f"\n❌ Error: {result['error']}")
        sys.exit(1)
    
    print(f"\n✅ Procesamiento completo:")
    print(f"  Total filas: {result['total_rows']}")
    print(f"  Válidas: {result['valid']}")
    print(f"  Inválidas: {result['invalid']}")
    print(f"  Geocodificadas: {result['geocoded']}")
    
    if args.dry_run:
        print(f"  Se insertarían: {result.get('would_insert', 0)} registros")
    else:
        print(f"  Insertadas/Actualizadas: {result.get('inserted', 0)}")
    
    if result['errors']:
        print(f"\n⚠️  Errores encontrados ({min(len(result['errors']), 10)} primeros):")
        for error in result['errors'][:10]:
            print(f"  - {error}")
        if len(result['errors']) > 10:
            print(f"  ... y {len(result['errors']) - 10} errores más")
    
    print(f"\n{'🔍 DRY RUN completado' if args.dry_run else '💾 Importación completada'}")


if __name__ == '__main__':
    main()
