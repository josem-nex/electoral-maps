"""Unit tests for personal_parser.parse_personal_file.

Uses an in-memory SQLite DB with minimal fixture data (a few puestos, one
departamento, one municipio) so tests run fast without any real DB.
"""
from __future__ import annotations

import io
from unittest.mock import MagicMock

import openpyxl
import pytest
from sqlalchemy import create_engine
from sqlalchemy.orm import sessionmaker
from sqlalchemy.pool import StaticPool

from app.database import Base
from app.db_models import (
    PuestoORM,
    TerritorioDepartamentoORM,
    TerritorioMunicipioORM,
    TerritorioZonaORM,
)
from app.personal_parser import parse_personal_file


# ── Fixtures ─────────────────────────────────────────────────────────────────

@pytest.fixture(scope="module")
def db_session():
    """In-memory DB with minimal territorial + puestos data."""
    engine = create_engine(
        "sqlite://",
        connect_args={"check_same_thread": False},
        poolclass=StaticPool,
    )
    Base.metadata.create_all(bind=engine)
    Session = sessionmaker(bind=engine)
    session = Session()

    # Zona
    session.add(TerritorioZonaORM(codigo="01", nombre="ZONA NOROCCIDENTAL"))
    # Departamento
    session.add(TerritorioDepartamentoORM(codigo="05", nombre="ANTIOQUIA", zona_codigo="01"))
    # Municipio
    session.add(TerritorioMunicipioORM(codigo="05001", nombre="MEDELLÍN", departamento_codigo="05"))
    session.add(TerritorioMunicipioORM(codigo="05002", nombre="ABEJORRAL", departamento_codigo="05"))

    # Puestos
    session.add(PuestoORM(
        codigo_puesto="050010101", departamento_codigo="05", municipio_codigo="05001",
        departamento="ANTIOQUIA", municipio="MEDELLÍN",
        puesto="IE LA ESPERANZA", latitud=6.25, longitud=-75.56,
    ))
    session.add(PuestoORM(
        codigo_puesto="050010102", departamento_codigo="05", municipio_codigo="05001",
        departamento="ANTIOQUIA", municipio="MEDELLÍN",
        puesto="IE LA CANDELARIA", latitud=6.25, longitud=-75.56,
    ))
    session.add(PuestoORM(
        codigo_puesto="050020101", departamento_codigo="05", municipio_codigo="05002",
        departamento="ANTIOQUIA", municipio="ABEJORRAL",
        puesto="SEDE CENTRAL ABEJORRAL", latitud=5.79, longitud=-75.43,
    ))
    session.commit()
    yield session
    session.close()


# ── Excel builders ────────────────────────────────────────────────────────────

def _make_jurados_xlsx(rows: list[list]) -> bytes:
    """Build an in-memory jurados xlsx with the standard headers."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Formulario"
    ws.append([
        "DEPARTAMENTO", "MUNICIPIO", "PUESTO", "CEDULA",
        "PRIMER NOMBRE", "SEGUNDO NOMBRE", "PRIMER APELLIDO", "SEGUNDO APELLIDO",
        "DIRECCION", "TELEFONO", "CELULAR", "CORREO", "NIVEL EDUCATIVO", "REFERENCIADO POR",
    ])
    for row in rows:
        ws.append(row)
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _make_testigos_xlsx(rows: list[list]) -> bytes:
    """Build an in-memory testigos xlsx with the standard headers."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Formulario"
    ws.append([
        "DEPARTAMENTO", "MUNICIPIO", "CEDULA", "PRIMER NOMBRE", "SEGUNDO NOMBRE",
        "PRIMER APELLIDO", "SEGUNDO APELLIDO", "CEDULA",
        "TELEFONO", "CORREO", "CELULAR", "COMUNA/LOCALIDAD/ZONA",
        "PUESTO DE VOTACION OPCION 1", "REFERENCIADO POR",
    ])
    for row in rows:
        ws.append(row)
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _make_jurados_con_codigo_xlsx(rows: list[list]) -> bytes:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Formulario"
    ws.append([
        "DEPARTAMENTO", "MUNICIPIO", "PUESTO", "CODIGO", "CEDULA",
        "PRIMER NOMBRE", "SEGUNDO NOMBRE", "PRIMER APELLIDO", "SEGUNDO APELLIDO",
        "DIRECCION", "TELEFONO", "CELULAR", "CORREO", "NIVEL EDUCATIVO", "REFERENCIADO POR",
    ])
    for row in rows:
        ws.append(row)
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


# ── Tests: tipo detection ─────────────────────────────────────────────────────

def test_detects_jurado_by_nivel_educativo(db_session):
    xlsx = _make_jurados_xlsx([
        ["ANTIOQUIA", "MEDELLÍN", "IE LA ESPERANZA", "12345678",
         "ANA", "", "GARCIA", "", "", "", "3001234567", "", "BACHILLER", ""],
    ])
    result = parse_personal_file(xlsx, "test.xlsx", db_session)
    assert result.tipo == "jurado"


def test_detects_testigo_by_puesto_col(db_session):
    xlsx = _make_testigos_xlsx([
        ["ANTIOQUIA", "MEDELLÍN", "11111111", "CARLOS", "",
         "LOPEZ", "", "11111111", "", "", "3009876543", "", "IE LA ESPERANZA", ""],
    ])
    result = parse_personal_file(xlsx, "test.xlsx", db_session)
    assert result.tipo == "testigo"


def test_tipo_override_forces_type(db_session):
    # File has NIVEL EDUCATIVO (jurado signal) but we force testigo — should fail gracefully
    # Actually override should be respected; here we use a file without distinguishing cols
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Formulario"
    ws.append(["DEPARTAMENTO", "MUNICIPIO", "PUESTO DE VOTACION OPCION 1", "CEDULA",
               "PRIMER NOMBRE", "PRIMER APELLIDO"])
    ws.append(["ANTIOQUIA", "MEDELLÍN", "IE LA ESPERANZA", "99999999", "JUAN", "DIAZ"])
    buf = io.BytesIO()
    wb.save(buf)
    result = parse_personal_file(buf.getvalue(), "test.xlsx", db_session, tipo_override="jurado")
    assert result.tipo == "jurado"


def test_raises_if_tipo_undetectable(db_session):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Formulario"
    ws.append(["DEPARTAMENTO", "MUNICIPIO", "CEDULA", "PRIMER NOMBRE"])
    ws.append(["ANTIOQUIA", "MEDELLÍN", "12345678", "ANA"])
    buf = io.BytesIO()
    wb.save(buf)
    with pytest.raises(ValueError, match="tipo de archivo"):
        parse_personal_file(buf.getvalue(), "test.xlsx", db_session)


# ── Tests: name resolution ────────────────────────────────────────────────────

def test_resolves_puesto_by_name(db_session):
    xlsx = _make_jurados_xlsx([
        ["ANTIOQUIA", "MEDELLÍN", "IE LA ESPERANZA", "12345678",
         "ANA", "", "GARCIA", "", "", "", "3001234567", "", "BACHILLER", ""],
    ])
    result = parse_personal_file(xlsx, "test.xlsx", db_session)
    assert result.omitidos == 0
    assert len(result.filas) == 1
    assert result.filas[0]["codigo_puesto"] == "050010101"


def test_normalizes_tildes_in_department(db_session):
    """'ANTIÓQUIA' (with tilde) should resolve to dep code 05."""
    xlsx = _make_jurados_xlsx([
        ["ANTIÓQUIA", "MEDELLÍN", "IE LA ESPERANZA", "22222222",
         "LUIS", "", "PEREZ", "", "", "", "", "", "TECNICO", ""],
    ])
    result = parse_personal_file(xlsx, "test.xlsx", db_session)
    assert result.omitidos == 0
    assert result.filas[0]["codigo_puesto"] == "050010101"


def test_unknown_puesto_generates_error_with_suggestion(db_session):
    """A puesto name close to a real one should include a suggestion."""
    xlsx = _make_jurados_xlsx([
        ["ANTIOQUIA", "MEDELLÍN", "IE LA ESPERANZ",  # missing last char
         "33333333", "PEDRO", "", "RAMIREZ", "", "", "", "", "", "BACHILLER", ""],
    ])
    result = parse_personal_file(xlsx, "test.xlsx", db_session)
    assert result.omitidos == 1
    assert len(result.errores) == 1
    # Suggestion should mention the real puesto name
    assert "Quiso decir" in result.errores[0]["razon"] or "esperanza" in result.errores[0]["razon"].lower()


def test_unknown_puesto_no_suggestion_when_too_different(db_session):
    """A totally different puesto name should not suggest anything."""
    xlsx = _make_jurados_xlsx([
        ["ANTIOQUIA", "MEDELLÍN", "ZZZZZZZ NADA",
         "44444444", "PEDRO", "", "RAMIREZ", "", "", "", "", "", "BACHILLER", ""],
    ])
    result = parse_personal_file(xlsx, "test.xlsx", db_session)
    assert result.omitidos == 1
    assert "Quiso decir" not in result.errores[0]["razon"]


def test_unknown_departamento_generates_error(db_session):
    xlsx = _make_jurados_xlsx([
        ["GUAJIRA", "RIOHACHA", "ALGUNA ESCUELA", "55555555",
         "ANA", "", "GARCIA", "", "", "", "", "", "BACHILLER", ""],
    ])
    result = parse_personal_file(xlsx, "test.xlsx", db_session)
    assert result.omitidos == 1
    assert "Departamento no encontrado" in result.errores[0]["razon"]


def test_unknown_municipio_generates_error(db_session):
    xlsx = _make_jurados_xlsx([
        ["ANTIOQUIA", "CIUDAD INVENTADA", "IE LA ESPERANZA", "66666666",
         "ANA", "", "GARCIA", "", "", "", "", "", "BACHILLER", ""],
    ])
    result = parse_personal_file(xlsx, "test.xlsx", db_session)
    assert result.omitidos == 1
    assert "Municipio no encontrado" in result.errores[0]["razon"]


def test_blank_rows_skipped_silently(db_session):
    xlsx = _make_jurados_xlsx([
        ["ANTIOQUIA", "MEDELLÍN", "IE LA ESPERANZA", "77777777",
         "ANA", "", "GARCIA", "", "", "", "", "", "BACHILLER", ""],
        ["", "", "", "", "", "", "", "", "", "", "", "", "", ""],  # blank
        ["", "", "", "", "", "", "", "", "", "", "", "", "", ""],  # blank
    ])
    result = parse_personal_file(xlsx, "test.xlsx", db_session)
    assert len(result.filas) == 1
    assert result.omitidos == 0


def test_missing_cedula_generates_error(db_session):
    xlsx = _make_jurados_xlsx([
        ["ANTIOQUIA", "MEDELLÍN", "IE LA ESPERANZA", "",  # empty cedula
         "ANA", "", "GARCIA", "", "", "", "", "", "BACHILLER", ""],
    ])
    result = parse_personal_file(xlsx, "test.xlsx", db_session)
    assert result.omitidos == 1
    assert "Cédula vacía" in result.errores[0]["razon"]


def test_partial_load_continues_after_errors(db_session):
    """Rows with errors should not abort processing of valid rows."""
    xlsx = _make_jurados_xlsx([
        ["ANTIOQUIA", "MEDELLÍN", "IE LA ESPERANZA", "88888881",
         "ANA", "", "GARCIA", "", "", "", "", "", "BACHILLER", ""],
        ["ANTIOQUIA", "MEDELLÍN", "PUESTO INVENTADO", "88888882",
         "LUIS", "", "PEREZ", "", "", "", "", "", "TECNICO", ""],
        ["ANTIOQUIA", "MEDELLÍN", "IE LA CANDELARIA", "88888883",
         "MARIO", "", "DIAZ", "", "", "", "", "", "UNIVERSITARIO", ""],
    ])
    result = parse_personal_file(xlsx, "test.xlsx", db_session)
    assert len(result.filas) == 2      # rows 1 and 3 resolved
    assert result.omitidos == 1        # row 2 failed


# ── Tests: CODIGO column (direct code path) ───────────────────────────────────

def test_codigo_column_resolves_directly(db_session):
    """When CODIGO column is present with a valid code, uses it directly."""
    xlsx = _make_jurados_con_codigo_xlsx([
        ["ANTIOQUIA", "MEDELLÍN", "NOMBRE INCORRECTO", "050010102",
         "12312312", "SOFIA", "", "VARGAS", "",
         "", "", "", "", "BACHILLER", ""],
    ])
    result = parse_personal_file(xlsx, "test.xlsx", db_session)
    assert result.omitidos == 0
    assert result.filas[0]["codigo_puesto"] == "050010102"


def test_codigo_column_leading_zero_preserved(db_session):
    """Codes with leading zeros must not lose them (xlsx numeric parsing)."""
    xlsx = _make_jurados_con_codigo_xlsx([
        ["ANTIOQUIA", "MEDELLÍN", "IE LA ESPERANZA", "050010101",
         "99912399", "JUAN", "", "MORA", "",
         "", "", "", "", "TECNICO", ""],
    ])
    result = parse_personal_file(xlsx, "test.xlsx", db_session)
    assert result.omitidos == 0
    assert result.filas[0]["codigo_puesto"] == "050010101"


def test_invalid_codigo_generates_error(db_session):
    """A CODIGO that doesn't exist in the DB should be reported as error."""
    xlsx = _make_jurados_con_codigo_xlsx([
        ["ANTIOQUIA", "MEDELLÍN", "IE LA ESPERANZA", "999999999",
         "11122233", "JUAN", "", "MORA", "",
         "", "", "", "", "TECNICO", ""],
    ])
    result = parse_personal_file(xlsx, "test.xlsx", db_session)
    assert result.omitidos == 1
    assert "Código de puesto no encontrado" in result.errores[0]["razon"]


def test_empty_codigo_falls_back_to_name(db_session):
    """If CODIGO column exists but cell is empty, fall back to name resolution."""
    xlsx = _make_jurados_con_codigo_xlsx([
        ["ANTIOQUIA", "MEDELLÍN", "IE LA ESPERANZA", "",  # empty CODIGO
         "55566677", "ELENA", "", "MORA", "",
         "", "", "", "", "TECNICO", ""],
    ])
    result = parse_personal_file(xlsx, "test.xlsx", db_session)
    # Name resolution should succeed
    assert result.omitidos == 0
    assert result.filas[0]["codigo_puesto"] == "050010101"


# ── Tests: field mapping ──────────────────────────────────────────────────────

def test_jurado_fields_mapped_correctly(db_session):
    xlsx = _make_jurados_xlsx([
        ["ANTIOQUIA", "MEDELLÍN", "IE LA ESPERANZA", "10203040",
         "CARLOS", "ANDRES", "MARTINEZ", "GOMEZ",
         "CRA 10 #20-30", "6011234567", "3101234567", "carlos@test.co",
         "UNIVERSITARIO", "PARTIDO X"],
    ])
    result = parse_personal_file(xlsx, "test.xlsx", db_session)
    f = result.filas[0]
    assert f["tipo"] == "jurado"
    assert f["cedula"] == "10203040"
    assert f["primer_nombre"] == "CARLOS"
    assert f["segundo_nombre"] == "ANDRES"
    assert f["primer_apellido"] == "MARTINEZ"
    assert f["segundo_apellido"] == "GOMEZ"
    assert f["direccion"] == "CRA 10 #20-30"
    assert f["nivel_educativo"] == "UNIVERSITARIO"
    assert f["referenciado_por"] == "PARTIDO X"
    assert f["codigo_puesto"] == "050010101"


def test_testigo_fields_mapped_correctly(db_session):
    xlsx = _make_testigos_xlsx([
        ["ANTIOQUIA", "MEDELLÍN", "50607080", "DIANA", "LUCIA",
         "HERRERA", "SILVA", "50607080",
         "6019876543", "diana@test.co", "3209876543", "CENTRO",
         "IE LA CANDELARIA", "NINGUNO"],
    ])
    result = parse_personal_file(xlsx, "test.xlsx", db_session)
    f = result.filas[0]
    assert f["tipo"] == "testigo"
    assert f["cedula"] == "50607080"
    assert f["primer_nombre"] == "DIANA"
    assert f["codigo_puesto"] == "050010102"
    assert f.get("direccion") is None     # testigos don't have direccion
    assert f.get("nivel_educativo") is None


def test_testigo_duplicate_cedula_col_uses_first(db_session):
    """Testigos template has CEDULA duplicated; parser uses first occurrence."""
    xlsx = _make_testigos_xlsx([
        ["ANTIOQUIA", "MEDELLÍN", "11223344", "ROSA", "",
         "QUINTERO", "", "DIFFERENT_VALUE",  # second CEDULA col with different value
         "", "", "3001112233", "", "IE LA ESPERANZA", ""],
    ])
    result = parse_personal_file(xlsx, "test.xlsx", db_session)
    assert result.filas[0]["cedula"] == "11223344"  # first CEDULA wins


# ── Tests: CSV support ────────────────────────────────────────────────────────

def test_csv_jurados_parsed(db_session):
    csv_content = (
        "DEPARTAMENTO,MUNICIPIO,PUESTO,CEDULA,PRIMER NOMBRE,SEGUNDO NOMBRE,"
        "PRIMER APELLIDO,SEGUNDO APELLIDO,DIRECCION,TELEFONO,CELULAR,CORREO,"
        "NIVEL EDUCATIVO,REFERENCIADO POR\n"
        "ANTIOQUIA,MEDELLÍN,IE LA ESPERANZA,99887766,LAURA,,TORRES,,,,3001119988,,TECNICO,\n"
    ).encode("utf-8")
    result = parse_personal_file(csv_content, "test.csv", db_session)
    assert result.tipo == "jurado"
    assert len(result.filas) == 1
    assert result.filas[0]["cedula"] == "99887766"
